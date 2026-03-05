# EXTREMO MASTER IA PRO - VERSÃO 7000 FINAL ESTÁVEL

import streamlit as st
import random
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="EXTREMO MASTER IA PRO", layout="centered")

st.title("🔥 EXTREMO MASTER IA PRO")
st.markdown("### Versão 7000 - Compatível com CSV e XLSX")

CONFIG = {
    "Mega-Sena": {"faixa": 60, "qtd": 6},
    "Lotofácil": {"faixa": 25, "qtd": 15},
    "Quina": {"faixa": 80, "qtd": 5},
    "Lotomania": {"faixa": 100, "qtd": 20},
}

def extrair_numeros(valor):
    numeros = []
    numero_atual = ""

    texto = str(valor)

    for c in texto:
        if c.isdigit():
            numero_atual += c
        else:
            if numero_atual:
                numero_int = int(numero_atual)
                if 1 <= numero_int <= 100:
                    numeros.append(numero_int)
                numero_atual = ""

    if numero_atual:
        numero_int = int(numero_atual)
        if 1 <= numero_int <= 100:
            numeros.append(numero_int)

    return numeros


def processar_csv(uploaded_file):
    df = pd.read_csv(uploaded_file, dtype=str, sep=None, engine="python")
    return df


def processar_xlsx(uploaded_file):
    bytes_data = uploaded_file.read()
    wb = load_workbook(filename=BytesIO(bytes_data), data_only=True)

    dados = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            dados.append(row)

    df = pd.DataFrame(dados)
    return df


def processar_arquivo(uploaded_file):
    try:
        if uploaded_file.name.endswith(".csv"):
            df = processar_csv(uploaded_file)
        else:
            df = processar_xlsx(uploaded_file)

        todos_numeros = []

        for coluna in df.columns:
            for valor in df[coluna]:
                numeros = extrair_numeros(valor)
                todos_numeros.extend(numeros)

        if not todos_numeros:
            st.warning("Nenhum número válido encontrado.")
            return None

        frequencia = (
            pd.Series(todos_numeros)
            .value_counts()
            .sort_index()
        )

        return frequencia

    except Exception as e:
        st.error(f"Erro ao processar arquivo: {e}")
        return None


def gerar_jogo(loteria):
    faixa = CONFIG[loteria]["faixa"]
    qtd = CONFIG[loteria]["qtd"]
    return sorted(random.sample(range(1, faixa + 1), qtd))


loteria = st.selectbox("Escolha a Loteria:", list(CONFIG.keys()))
uploaded_file = st.file_uploader("Envie o arquivo CSV ou XLSX", type=["xlsx", "csv"])

if uploaded_file:
    st.success("Arquivo carregado com sucesso!")

    frequencia = processar_arquivo(uploaded_file)

    if frequencia is not None:
        st.subheader("Frequência dos números encontrados:")
        st.dataframe(frequencia)

        if st.button("Gerar Jogo Inteligente"):
            jogo = gerar_jogo(loteria)
            st.success(f"Jogo gerado: {jogo}")
